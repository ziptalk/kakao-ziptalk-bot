# -*- coding: utf-8 -*-
 
from io import DEFAULT_BUFFER_SIZE
import os
from flask import Flask, request, jsonify
import json
import logging
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from urllib.parse import urlencode, quote_plus
from urllib.request import Request, urlopen
from lxml import html
import bs4
import pandas as pd
from datetime import datetime, timedelta
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
import sys

from googletrans import Translator
from predict import predict_unseen_data

 
 
 
app = Flask(__name__)


########################

try:
    cred = credentials.Certificate(
        'ziptalk-chatbot-firebase-adminsdk-kz477-4cadf62941.json')
    firebase_admin.initialize_app(cred, {
        'projectId': 'ziptalk-chatbot',
    })
    print("파베 연결 완료")


    db = firestore.client()
    print("디비 연결 완료")

    # docs = db.collection(u'subscription_info').where(u'realtime_info.date', u'==', '2021-01-18').stream()
    docs = db.collection(u'subscription_info').stream()
    print("파베 불러옴")
except:
    print("firebase init error")
    pass


########################

def get_subscription_list(bbs_tl='', bbs_dtl_cts='', category='', detail_category=''):
    url = 'http://apis.data.go.kr/B552555/lhNoticeInfo/getNoticeInfo'
    # service_key = 'PdWFVj9WjaMQ7Qmoamq2n1f81jXwnfinEaCxcbGTtjmlmpwPcfEsQkky9Cdgz6J%2BtWUeGpU5BaVi6fZsgnL9qw%3D%3D'
    service_key = 'PdWFVj9WjaMQ7Qmoamq2n1f81jXwnfinEaCxcbGTtjmlmpwPcfEsQkky9Cdgz6J+tWUeGpU5BaVi6fZsgnL9qw=='

    
    today = datetime.today()
    sch_ed_dt = today.strftime("%Y-%m-%d")

    months_ago = datetime.today() - timedelta(365) # 최근 일년 단위로 검색
    sch_st_dt = months_ago.strftime("%Y-%m-%d")

    pg_sz = '10'    # 한 페이지 결과 수
    page = '1'      # 페이지 번호

    upp_ais_tp_cd = '99'        # 상위유형코드
    ais_tp_cd = ''              # 유형코드
    ais_tp_cd_int = ''          # 유형코드 추가
    ais_tp_cd_int2 = ''         # 유형코드 추가2
    ais_tp_cd_int3 = ''         # 유형코드 추가3

    if category == '토지':
        upp_ais_tp_cd = '01'
        if detail_category == '주택용지':
            ais_tp_cd = '02'
        elif detail_category == '상가용지':
            ais_tp_cd = '03'
        elif detail_category == '산업시설용지':
            ais_tp_cd = '04'
        elif detail_category == '기타용지':
            ais_tp_cd = '28'
    elif category == '분양주택':
        upp_ais_tp_cd = '05'
    elif category == '신혼희망타운':
        upp_ais_tp_cd = '39'
    elif category == '임대주택':
        upp_ais_tp_cd = '06'
        if detail_category == '국민임대':
            ais_tp_cd = '07'
        elif detail_category == '공공임대':
            ais_tp_cd = '08'
        elif detail_category == '영구임대':
            ais_tp_cd = '09'
        elif detail_category == '행복주택':
            ais_tp_cd = '10'
        elif detail_category == '장기전세':
            ais_tp_cd = '11'
        elif detail_category == '신축다세대':
            ais_tp_cd = '12'
    elif category == '매입임대' or category == '전세임대':
        upp_ais_tp_cd = '13'
        if detail_category == '매입임대':
            ais_tp_cd_int2 = '26'
        elif detail_category == '전세임대':
            ais_tp_cd_int3 = '17'
        elif detail_category == '집주인임대':
            ais_tp_cd_int = '36'
    elif category == '상가':
        upp_ais_tp_cd = '22'
        if detail_category == '분양(구)임대상가(입찰)':
            ais_tp_cd = '22'
        elif detail_category == '임대상가(입찰)':
            ais_tp_cd = '43'
        elif detail_category == '임대상가(공모심사)':
            ais_tp_cd = '38'
        elif detail_category == '임대상가(추첨)':
            ais_tp_cd = '24'
    elif category != '': # 위의 경우들이 아니면서 category가 비어있지 않으면 '기타'로 처리
        upp_ais_tp_cd = '99'
    else:
        # upp_ais_tp_cd = ''  # 카테고리 지정이 이도저도 아니어서 오류인 경우
        return '카테고리 지정이 잘못 되었습니다.'


    try:
        queryParams = '?' + urlencode({quote_plus('ServiceKey'): service_key, quote_plus('PG_SZ'): pg_sz, quote_plus('SCH_ST_DT'): sch_st_dt, quote_plus('SCH_ED_DT'): sch_ed_dt, quote_plus('BBS_TL'): bbs_tl, quote_plus(
            'BBS_DTL_CTS'): bbs_dtl_cts, quote_plus('UPP_AIS_TP_CD'): upp_ais_tp_cd, quote_plus('AIS_TP_CD'): ais_tp_cd, quote_plus('AIS_TP_CD_INT'): ais_tp_cd_int, quote_plus('AIS_TP_CD_INT2'): ais_tp_cd_int2, quote_plus('AIS_TP_CD_INT3'): ais_tp_cd_int3, quote_plus('PAGE'): page})

        request = Request(url + queryParams)
        request.get_method = lambda: 'GET'
        response_body = urlopen(request).read()

        result_body = response_body.decode('utf-8')

        data = json.loads(result_body)
        data_list = data[1]
        real_data = data_list['dsList']

        # print(real_data)
        # df = pd.DataFrame(real_data)

        url = 'http://apis.data.go.kr/B552555/lhNoticeDtlInfo/getNoticeDtlInfo'
        service_key = 'PdWFVj9WjaMQ7Qmoamq2n1f81jXwnfinEaCxcbGTtjmlmpwPcfEsQkky9Cdgz6J+tWUeGpU5BaVi6fZsgnL9qw=='
        queryParams = '?' + urlencode({ quote_plus('ServiceKey') : service_key, quote_plus('CCR_CNNT_SYS_DS_CD') : '01', quote_plus('BBS_SN') : '90000001', quote_plus('') : '' })


        ccr_cnnt_sys_ds_cd = ''
        bbs_sn = ''

        result_comment = ''

        result_comment = result_comment + '검색 시작일 : ' + sch_st_dt + '\n'
        result_comment = result_comment + '검색 종료일 : ' + sch_ed_dt + '\n\n▼▼▼▼ 검색 결과 ▼▼▼▼\n\n'


        # for row in range(0, len(real_data)):
        for row in range(0, 2):
            print(real_data[row]['CCR_CNNT_SYS_DS_CD'])
            print(real_data[row]['BBS_SN'])
            ccr_cnnt_sys_ds_cd = real_data[row]['CCR_CNNT_SYS_DS_CD']
            bbs_sn = real_data[row]['BBS_SN']

            queryParams = '?' + urlencode({ quote_plus('ServiceKey') : service_key, quote_plus('CCR_CNNT_SYS_DS_CD') : ccr_cnnt_sys_ds_cd, quote_plus('BBS_SN') : bbs_sn, quote_plus('') : '' })

            request = Request(url + queryParams)
            request.get_method = lambda: 'GET'
            response_body = urlopen(request).read()

            result_body = response_body.decode('utf-8')

            try:
                data = json.loads(result_body)
                data_list = data[1]
                real_data2 = data_list['dsBbsInfo']
                download_data = data_list['dsBbsAhflInfo']
                detail_list = real_data2[0]
                download_detail_list = download_data[0]
                # print(real_data2)
                try:
                    # print(detail_list)
                    # print(detail_list['AHFL_URL'])   # 다운로드 링크만 출력하기
                    # print(download_detail_list['AHFL_URL'])
                    
                    result_comment = result_comment + '제목 : ' + detail_list['BBS_TL'] + '\n' 
                    result_comment = result_comment + '게시일 : ' + detail_list['BBS_WOU_DTTM'] + '\n'
                    result_comment = result_comment + '담당부서 : ' + detail_list['DEP_NM'] + '\n'
                    result_comment = result_comment + '내용 : ' + detail_list['BBS_DTL_CTS'] + '\n'
                    try:
                        result_comment = result_comment + '파일명 : ' + download_detail_list['CMN_AHFL_NM'] + '\n'
                        result_comment = result_comment + '파일링크 : ' + download_detail_list['AHFL_URL'] + '\n======================\n'
                    except:
                        pass
                except:
                    # print('다운로드 링크 없음')
                    pass
            except:
                pass
        
    except:
        return "검색 결과가 없습니다."
        # return upp_ais_tp_cd + ' | ' + ais_tp_cd + ' | ' + ais_tp_cd_int + ' | ' + ais_tp_cd_int2 + ' | ' + ais_tp_cd_int3
        

    return result_comment


########################

def get_act_apt_list(area_code, year_mon):
    # area_code = '11410'
    # year_mon = '201912'
    print(area_code)
    print(year_mon)
    try:
        # url = "https://korbillgates.tistory.com"
        

        print("여긴와?0")
        # url = 'http://openapi.molit.go.kr/OpenAPI_ToolInstallPackage/service/rest/RTMSOBJSvc/getRTMSDataSvcAptTradeDev'
        # url = 'http://apis.data.go.kr/B552555/lhNoticeDtlInfo/getNoticeDtlInfo'
        url = 'http://openapi.molit.go.kr:8081/OpenAPI_ToolInstallPackage/service/rest/RTMSOBJSvc/getRTMSDataSvcAptTrade'
        service_key = 'PdWFVj9WjaMQ7Qmoamq2n1f81jXwnfinEaCxcbGTtjmlmpwPcfEsQkky9Cdgz6J+tWUeGpU5BaVi6fZsgnL9qw=='  # 서비스 인증키
        # service_key = 'OBnmXFNjkhuaUowPaTrGaDigDegJTAKSSjjO2kHz/RDSkSjHmd/V/CXhvd6E9NRuNnPwzuW3ij+biAdy52aTEQ=='
        
        res = urlopen(url)
        print(res.status)  ## 200
        
        
        # queryParams = '?' + urlencode(
        #     {
        #         quote_plus('ServiceKey'): service_key,
        #         quote_plus('LAWD_CD'): area_code,
        #         quote_plus('DEAL_YMD'): year_mon
        #     }
        # )

        # print(year_mon)

        # request = Request(url + queryParams)
        # request.get_method = lambda: 'GET'
        # # response = urlopen(request)
        # # response = requests.get(url, data=queryParams)
        # response_body = urlopen(request).read()

        queryParams = '?' + urlencode(
            {
                quote_plus('ServiceKey'): service_key,
                quote_plus('LAWD_CD'): area_code,
                quote_plus('DEAL_YMD'): year_mon
            }
        )

        request = Request(url + queryParams)
        request.get_method = lambda: 'GET'
        response_body = urlopen(request).read()
        print("여기는 옵니까...")
        # response_body = response.read()
        print("여긴와?4")

        result_body = response_body.decode('utf-8')
        print("get_act_apt_list 함수 끝까지 도달스")
        print(result_body)

        return result_body

    except:
        return "get_act_apt_list 함수 오류 발생"



########################

def get_act_apt_parsing_pd(result_body, dongname):
    print("dongname")
    print(dongname)
    try:
        try:
            xmlobj = bs4.BeautifulSoup(result_body, 'lxml-xml')
        except:
            return "bs4 오류"

        try:
            rows = xmlobj.findAll('item')
        except:
            return "xmlobj 오류"
        # print(rows[0])
        columns = rows[0].find_all()
        # print(columns)

        print(columns[0].name)
        print(columns[0].text)

        rowList = []
        nameList = []
        columnList = []
        result = ''

        rowsLen = len(rows)
        # for i in range(0, rowsLen):
        try:
            for i in range(1, rowsLen):
                columns = rows[i].find_all()
                columnsLen = len(columns)

                for j in range(0, columnsLen):

                    if i == 0:
                        nameList.append(columns[j].name)
                    else:
                        # 동이름이 같으면! columns[3] == (' ' + dongname)
                        if columns[3].text == (' ' + dongname):
                            result = result + columns[j].name + \
                                ' : ' + columns[j].text + '\n'

                    eachColumn = columns[j].text
                    columnList.append(eachColumn)

                if columns[3].text == (' ' + dongname):
                    result = result + '\n---------------------\n'
                    print(result)

                rowList.append(columnList)
                columnList = []
            # result = '뭐가 문젠데 싯팔...'

            # 거래금액  건축년도     년   법정동  아파트   월   일   전용면적      지번   지역코드   층
            # result = pd.DataFrame(rowList, columns=nameList)
            print(result)

            return result
        except:
            return "result 오류"

    except:
        return "get_act_apt_parsing_pd 함수 오류 발생"



########################

def get_weather(where):
    print("여긴 오긴 하냐?")
    weather = ""
    url = "https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query={}+날씨".format(
        where)
    # print(url)
    r = requests.get(url)
    # print(r)
    bs = BeautifulSoup(r.text, "lxml")
    # print(bs)
    weather_info = bs.select("div.today_area > div.main_info") #왜 여기서 null을 반환하지??
    print("여긴 오긴 하냐?22")
    # print(where)
    # print(weather_info)


    if len(weather_info) > 0:
        temperature = bs.select("span.todaytemp")
        cast_text = bs.select("p.cast_txt")
        indicator = bs.select("span.indicator")
        # print("여긴와??????")

        if len(temperature) > 0 and len(cast_text) > 0 and len(indicator) > 0:
            temperature = temperature[0].text.strip()
            indicator = indicator[0].text.strip()
            txt = cast_text[0].text.strip()

            print(temperature, indicator, txt)

            weather = "{}도\r\n{}\r\n{}".format(temperature, indicator, txt)

            return weather



########################


def api_AptList2(search_code_value, date_value, dongname_value):
    if date_value.isdigit():
        total_data = get_act_apt_list(
            area_code=search_code_value, year_mon=date_value)
        data_table = get_act_apt_parsing_pd(
            result_body=total_data, dongname=dongname_value)
        return (data_table, "Success")
    else:
        return (None, "year_mon date error")


########################

def api_AptList(area_command):
    print(area_command)
    area_command = area_command.split(" ")

    print(area_command)

    try:
        cityname = area_command[0]
        guname = area_command[1]
        date = area_command[-1]
        print(cityname)
        print(guname)
        print(date)
        dongname = area_command[-2]

        print("동이름 확인하기")
        print(dongname)

        confirm = '1.' + cityname + '2.' + guname + '3.' + date + '4.' + dongname
        # return (None, confirm)

        # try:
        wb = load_workbook(filename='dongcode_20180703.xlsx')
        sheet = wb['Sheet1']

        dongcode = " "

        for i in range(1, 3000):
            # for i in range(1, 18858):
            if sheet[i][2].value == cityname and (guname in sheet[i][3].value):
                dongcode = sheet[i][1].value
                break

        if dongcode == " ":
            return (None, "dongcode error")

        else:
            # 동코드가 1111010100 이런 형식이므로 앞에 11110 만 가져오도록 인덱싱.
            search_code = dongcode[0:5]
            print("검색코드" + search_code)

            return api_AptList2(search_code, date, dongname)

    except:
        return (None, "명령어 인식 분할부터 문제 발생")


########################



########################
 
@app.route('/keyboard')
def Keyboard():
 
    dataSend = {
        "type" : "buttons",
        "buttons" : ["시작하기", "도움말"]
    }
 
    return jsonify(dataSend)
 
 
# @app.route('/message')
@app.route('/message', methods=['POST'])
def Message():
    
    content = request.get_json()
    user_id = content['intent']
    block_name = user_id['name']

    user_id = user_id['id']
    content = content['userRequest']

    user_id2 = content['user']
    user_id2 = user_id2['id']
    
    content = content['utterance']

    print("user id : " + user_id2)
    print("block name : " + block_name)
    print("What did he/she say? : " + content)

    today = datetime.today()
    yyyy_mm_dd = today.strftime("%Y-%m-%d")

    is_question = False

    db_user = firestore.client()
    print("디비 연결 완료")

    docs_user = db_user.collection(u'user_record').document(user_id2)


    # docs = db.collection(u'subscription_info').where(u'realtime_info.date', u'==', '2021-01-18').stream()
    try:
        print("여긴와?")
        docs_user.set({
            u'date' : yyyy_mm_dd,
            u'user_id' : user_id2,
            u'block_name' : block_name,
            u'comment' : content
        }, merge=True)
    except:
        print("adding data to db is failed.")
        # docs_user = db_user.collection(u'user_record').document()
    
    print("파베 불러옴")

    # content = "/날씨 남가좌동"

    text = ""

    if content == 'start':
        text = """안녕하세요! ziptok 챗봇을 이용해주셔서 감사합니다! 챗봇 사용법은 아래 설명을 참고해주세요.
        
    질문하기
    ex ) 질문하기 부동산 계약할 때 사기 당한 거 같은데 도와주세요 ㅠㅠ
    ex ) 질문하기 연말정산을 하려고 하는데 처음이라 잘 모르겠어요. 어떻게 시작해야하나요?
    -> 앞에 '질문하기'를 붙이고 자유롭게 부동산과 관련된 질문을 해주시면 질문 카테고리에 맞는 전문가 연결과 비슷한 사례들을 보여주는 명령어입니다.
    
    오늘 청약
    ex ) 오늘 청약
    ex ) 내일 청약
    ex ) 모레 청약
    -> 해당 일자 청약 일정 정보를 알려주는 명령어입니다.
    
    아파트실거래가 <시 이름> <구 이름> <동 이름> <년월 6자리>
    ex ) 아파트실거래가 서울특별시 종로구 사직동 202010 

    청약 <카테고리> <세부카테고리>
    ex ) 청약 임대주택 행복주택

    ** 카테고리 안내 **
    토지 -> 주택용지, 상가용지, 산업시설용지, 기타용지
    분양주택
    신혼희망타운
    임대주택 -> 국민임대, 공공임대, 영구임대, 행복주택, 장기전세, 신축다세대
    매입임대 -> 매입임대, 전세임대, 집주인임대
    상가 -> 분양(구)임대상가(입찰), 임대상가(입찰), 임대상가(공모심사), 임대상가(추첨)
    
    날씨 <지역명>
    ex ) 날씨 남가좌동
        """

    else:
        str_message = content
        # if str_message[0:1] == "":
        if True:
            args = str_message.split(" ")
            command = args[0]
            del args[0]

            if command == "날씨":
                w = " ".join(args)
                text = get_weather(w)
            
            elif command == "맞아요":
                user_prev = firestore.client()

                prev_data = user_prev.collection(u'user_record').document(user_id2).to_dict()
                question_tmp = prev_data['comment']
                block_tmp = prev_data['block_name']
                try:
                    category_tmp = prev_data['category']
                except:
                    pass

                good_data = user_prev.collection(u'good_data').document()
                if block_tmp == "폴백 블록":
                    try:
                        good_data.set({
                            "category": category_tmp,
                            "question": question_tmp
                        })
                        print("success!")
                    except:
                        print("category가 없나봅니다.")

                text = "소중한 정보 감사합니다. :)"

            elif command == "아니에요":
                text = "소중한 정보 감사합니다. :)"

            elif command == "help" or command == "도움말":
                text = """ziptok 챗봇을 이용해주셔서 감사합니다! 아래 명령어를 참고해주세요.

        질문하기
        ex ) 질문하기 부동산 계약할 때 사기 당한 거 같은데 도와주세요 ㅠㅠ
        ex ) 질문하기 연말정산을 하려고 하는데 처음이라 잘 모르겠어요. 어떻게 시작해야하나요?
        -> 앞에 '질문하기'를 붙이고 자유롭게 부동산과 관련된 질문을 해주시면 질문 카테고리에 맞는 전문가 연결과 비슷한 사례들을 보여주는 명령어입니다.

        오늘 청약
        ex ) 오늘 청약
        ex ) 내일 청약
        ex ) 모레 청약
        -> 해당 일자 청약 일정 정보를 알려주는 명령어입니다.
        
        아파트실거래가 <시 이름> <구 이름> <동 이름> <년월 6자리>
        ex ) 아파트실거래가 서울특별시 종로구 사직동 202010 

        청약 <카테고리> <세부카테고리>
        ex ) 청약 임대주택 행복주택

        ** 카테고리 안내 **
        토지 -> 주택용지, 상가용지, 산업시설용지, 기타용지
        분양주택
        신혼희망타운
        임대주택 -> 국민임대, 공공임대, 영구임대, 행복주택, 장기전세, 신축다세대
        매입임대 -> 매입임대, 전세임대, 집주인임대
        상가 -> 분양(구)임대상가(입찰), 임대상가(입찰), 임대상가(공모심사), 임대상가(추첨)
        
        날씨 <지역명>
        ex ) 날씨 남가좌동"""

            
            elif command == "오늘" or command == "내일" or command == "모레":
                # today = datetime.today()
                # sub_date = today.strftime("%Y-%m-%d")

                try:
                    # if(db == None):
                    

                    text = ""

                    # sub_date = "2021-03-29"
                    today_date = datetime.today()
                    sub_date = today_date.strftime("%Y-%m-%d")

                    if(command == "내일"):
                        tomorrow = today_date + timedelta(days=1)
                        sub_date = tomorrow.strftime("%Y-%m-%d")
                    
                    elif (command == "모레"):
                        da_tomorrow = today_date + timedelta(days=2)
                        sub_date = da_tomorrow.strftime("%Y-%m-%d")

                    db = firestore.client()
                    print("디비 연결 완료")

                    # docs = db.collection(u'subscription_info').where(u'realtime_info.date', u'==', '2021-01-18').stream()
                    docs = db.collection(u'subscription_info').stream()
                    print("파베 불러옴")
                     
                    for doc in docs:
                        temp = doc.to_dict()

                        # print(sub_date)

                        if(temp["realtime_info"]["date"] == sub_date):
                            try:
                                text = text + "날짜 : " + temp["realtime_info"]["date"] + "\n"
                                text = text + "▼▼▼ 아파트정보 ▼▼▼" + "\n"
                                text = text + "아파트명 : " + temp["realtime_info"]["apt_info"]["apt_name"] + "\n"
                                text = text + "공급위치 : " + temp["realtime_info"]["apt_info"]["address"] + "\n"
                                text = text + "공급규모 : " + temp["realtime_info"]["apt_info"]["sup_size"] + "\n"
                                text = text + "문의처 : " + temp["realtime_info"]["apt_info"]["tel"].replace("\n", "") + "\n"
                                text = text + "▼▼▼ 청약일정 ▼▼▼" + "\n"
                                text = text + "모집공고일 : " + temp["realtime_info"]["sub_sch"]["ann_date"] + "\n"
                                for sch in range(0, 3):
                                    if (temp["realtime_info"]["sub_sch"]["sub_rec"][sch]["class_name"] != ""):
                                        text = text + "구분명 : " + temp["realtime_info"]["sub_sch"]["sub_rec"][sch]["class_name"] + "\n"
                                        text = text + "해당지역 접수일 : " + temp["realtime_info"]["sub_sch"]["sub_rec"][sch]["local_date"] + "\n"
                                        text = text + "기타지역 접수일 : " + temp["realtime_info"]["sub_sch"]["sub_rec"][sch]["other_date"] + "\n"
                                        text = text + "접수장소 : " + temp["realtime_info"]["sub_sch"]["sub_rec"][sch]["recept_place"] + "\n"
                                    elif (sch == 0 and temp["realtime_info"]["sub_rec"][sch]["class_name"] == ""):
                                        text = text + "접수일 : " + temp["realtime_info"]["sub_sch"]["sub_rec"][sch]["local_date"] + "\n"
                                text = text + "당첨자 발표일 : " + temp["realtime_info"]["sub_sch"]["winner_date"].replace("\n", "").replace("\t", "") + "\n"
                                text = text + "계약일 : " + temp["realtime_info"]["sub_sch"]["contract_date"] + "\n"
                                text = text + "=================\n"
                            except:
                                pass
                    
                    if text == "":
                        text = "당일 접수 일정은 없습니다."

                except:
                    text = "오늘청약 명령어 에러"
                    text = sys.exc_info()[0]



            elif command == "아파트실거래가":
                w = " ".join(args)
                error_code = ""
                # try:
                #     property_table, error_code = api_AptList(w)
                # except:
                #     error_code = "api_AptList 함수 처리 과정에서 error 발생"

                # if error_code != "Success":
                #     text = error_code
                # else:
                #     # property_table = property_table.to_string()
                #     text = property_table
                area_command = w.split(" ")

                # print(area_command)

                cityname = area_command[0]
                guname = area_command[1]
                date = area_command[-1]
                
                dongname = area_command[-2]

                confirm = '1.' + cityname + '2.' + guname + '3.' + date + '4.' + dongname
                # return (None, confirm)

                # try:
                wb = load_workbook(filename='dongcode_20180703_real.xlsx')
                sheet = wb['Sheet1']

                dongcode = " "

                for i in range(1, 230):
                    # for i in range(1, 18858):

                    if (cityname[0:2] in sheet[i][2].value) and (guname in sheet[i][3].value):
                        dongcode = sheet[i][1].value
                        break

                if dongcode == " ":
                    # return (None, "dongcode error")
                    text = "dongcode error"

                else:
                    # 동코드가 1111010100 이런 형식이므로 앞에 11110 만 가져오도록 인덱싱.
                    search_code = dongcode[0:5]
                    print("검색코드" + search_code)

                    if date.isdigit():
                        try:
                            test_body = get_act_apt_list(int(search_code), int(date))
                            # test_body = get_act_apt_list('11110', '202010')
                            if (test_body != "get_act_apt_list 함수 오류 발생") and (test_body != ''):
                                print("get in!")
                                test_result = get_act_apt_parsing_pd(test_body, dongname)
                                print(test_result)

                                if isinstance(test_result, str):
                                    text = test_result
                                else:
                                    text = test_result.to_string()
                            else:
                                text = test_body
                        
                        except:
                            text = search_code
                        # text = search_code
                    else:
                        text = "year_mon date error"
                        # return (None, "year_mon date error")



            #     # test_body = get_act_apt_list(11110, 202010)
            #     # test_body = get_act_apt_list(int(serach_code), int(date))

            elif command == "질문하기":
                is_question = True

                w = " ".join(args) # 사용자가 질문한 내용
                error_code = ""
                
                test_file = './data/small_samples_property.json'
                test_examples = json.loads(open(test_file).read())

                test_examples[0]['category'] = "법률" # 비워놓으면 안돼서 그냥 아무거나로 초기화
                # print(type(kor_category))
                # print(kor_category)
                test_examples[0]['question'] = w

                print(test_examples[0]['category'])
                print(test_examples[0]['question'])

                result_examples = predict_unseen_data(test_examples[0]['category'], test_examples[0]['question']) #predict.py에서 가져온 모듈 돌려~

                print(result_examples)

                # result_file = './data/small_samples_prediction.json'
                # result_examples = json.loads(open(result_file).read())
                
                translator = Translator(service_urls=['translate.googleapis.com'])

                result_eng = result_examples[0]['new_category']

                result_kor = translator.translate(result_eng, dest='ko')

                result_tmp = result_kor.text

                if(result_tmp == "월간 간행물"):
                    result_tmp = "월세"
                elif(result_tmp == "판매, 구독"):
                    result_tmp = "분양, 청약"
                elif(result_tmp == "세"):
                    result_tmp = "세무"
                elif(result_tmp == "거래"):
                    result_tmp = "매매"
                elif(result_tmp == "관리"):
                    result_tmp = "행정"
                elif(result_tmp == "법"):
                    result_tmp = "법률"
                elif(result_tmp == "차관"):
                    result_tmp = "대출"

                ######################

                
                url = "https://kin.naver.com/search/list.nhn?query={}".format(
                    w)
                # print(url)
                r = requests.get(url)
                # print(r)
                bs = BeautifulSoup(r.text, "lxml")
                # print(bs)
                link_info1 = bs.select("#s_content > div.section > ul > li:nth-child(1) > dl > dt > a")
                # print(link_info)
                # print(link_info1[0])
                href1 = link_info1[0].attrs['href']

                url1 = href1
                r1 = requests.get(url1)
                bs1 = BeautifulSoup(r1.text, "lxml")
                title1 = bs1.select("#content > div.question-content > div > div.c-heading._questionContentsArea.c-heading--default-old > div.c-heading__title > div > div.title")
                question1 = bs1.select('#content > div.question-content > div > div.c-heading._questionContentsArea.c-heading--default-old > div.c-heading__content')
                answer1 = bs1.select('#answer_1 > div._endContents.c-heading-answer__content > div._endContentsText.c-heading-answer__content-user')

                try:
                    result_title1 = title1[0].text.strip()
                except:
                    result_title1 = ''

                try:
                    result_question1 = question1[0].text.strip()
                except:
                    result_question1 = ''

                try:
                    result_answer1 = answer1[0].text.strip()
                except:
                    result_answer1 = ''

                

                link_info2 = bs.select("#s_content > div.section > ul > li:nth-child(2) > dl > dt > a")
                href2 = link_info2[0].attrs['href']

                url2 = href2
                r2 = requests.get(url2)
                bs2 = BeautifulSoup(r2.text, "lxml")
                title2 = bs2.select("#content > div.question-content > div > div.c-heading._questionContentsArea.c-heading--default-old > div.c-heading__title > div > div.title")
                question2 = bs2.select('#content > div.question-content > div > div.c-heading._questionContentsArea.c-heading--default-old > div.c-heading__content')
                answer2 = bs2.select('#answer_1 > div._endContents.c-heading-answer__content > div._endContentsText.c-heading-answer__content-user')

                try:
                    result_title2 = title2[0].text.strip()
                except:
                    result_title2 = ''

                try:
                    result_question2 = question2[0].text.strip()
                except:
                    result_question2 = ''

                try:
                    result_answer2 = answer2[0].text.strip()
                except:
                    result_answer2 = ''

                link_info3 = bs.select("#s_content > div.section > ul > li:nth-child(3) > dl > dt > a")
                href3 = link_info3[0].attrs['href']

                url3 = href3
                r3 = requests.get(url3)
                bs3 = BeautifulSoup(r3.text, "lxml")
                title3 = bs3.select("#content > div.question-content > div > div.c-heading._questionContentsArea.c-heading--default-old > div.c-heading__title > div > div.title")
                question3 = bs3.select('#content > div.question-content > div > div.c-heading._questionContentsArea.c-heading--default-old > div.c-heading__content')
                answer3 = bs3.select('#answer_1 > div._endContents.c-heading-answer__content > div._endContentsText.c-heading-answer__content-user')

                try:
                    result_title3 = title3[0].text.strip()
                except:
                    result_title3 = ''

                try:
                    result_question3 = question3[0].text.strip()
                except:
                    result_question3 = ''

                try:
                    result_answer3 = answer3[0].text.strip()
                except:
                    result_answer3 = ''

                docs_user.set({
                    u'category' : result_tmp,
                }, merge=True)

                result1 = '질문제목 : ' + result_title1 + '\n' + '질문내용 : ' +result_question1 + '\n' + '답변내용 : ' + result_answer1 + '\n'
                result2 = '질문제목 : ' + result_title2 + '\n' + '질문내용 : ' +result_question2 + '\n' + '답변내용 : ' + result_answer2 + '\n'
                result3 = '질문제목 : ' + result_title3 + '\n' + '질문내용 : ' +result_question3 + '\n' + '답변내용 : ' + result_answer3 + '\n'

                text = result_tmp + "에 관련한 질문이네요! 해당 전문가와 연결해드릴까요?" + "\n\n다음은 가장 유사한 질문들입니다.\n\n" + result1 + '\n' + result2 + '\n' + result3 + '\n'

                dataSend = {
                    "version": "2.0",
                    "template": {"outputs": [{"simpleText": {"text": result_tmp + "에 관련한 질문이네요! 해당 전문가와 연결해드릴까요?"}}],
                                "quickReplies": [{"label": "good", "action": "message", "messageText": "👍 맞아요"},
                                                {"label": "bad", "action": "message", "messageText": "👎 아니에요"},
                                                ]
                                },
                    "message": {
                        # "text": unicode(text, "utf-8")
                        "text": result_tmp + "에 관련한 질문이네요! 해당 전문가와 연결해드릴까요?",
                        "text2": result_answer1,
                        "text3": result_answer2,
                        "text4": result_answer3,
                        "url": "https://ziptalk.imweb.me/"
                    }
                }

            elif command == "청약":
                w = " ".join(args)
                error_code = ""
                detail_command = w.split(" ")

                # print(area_command)

                category_ = detail_command[0]
                try:
                    detailCategory = detail_command[1]
                except:
                    pass

                try:
                    text = get_subscription_list(bbs_tl='', bbs_dtl_cts='', category=category_, detail_category=detailCategory)
                except:
                    text = "요청하신 정보를 불러올 수 없습니다."


            else:
                text = """잘못된 명령어 형식입니다. 다시 확인해주세요.
                도움이 필요하시다면 "help" 혹은 "도움말" 명령어를 이용해주세요!"""

        else:
            text = """잘못된 명령어 형식입니다. 다시 확인해주세요.
            도움이 필요하시다면 "help" 혹은 "도움말" 명령어를 이용해주세요!"""

 
    # if content == "시작하기":
    #     dataSend = {
    #         "message": {
    #             "text": "아직 개발중이라 대답을 잘 못해도 이해해줘^^;"
    #         }
    #     }
    # elif content == "도움말":
    #     dataSend = {
    #         "message": {
    #             "text": "이제 곧 정식 버전이 출시될거야. 조금만 기다려~~~"
    #         }
    #     }
    # elif "안녕" in content:
    #     dataSend = {
    #         "message": {
    #             "text": "안녕~~ 반가워 ㅎㅎ"
    #         }
    #     }
    # else:
    #     dataSend = {
    #         "message": {
    #             "text": "나랑 놀자 ㅋㅋㅋ"
    #         }
    #     }

    if (is_question == False):
        dataSend = {
            "message": {
                # "text": unicode(text, "utf-8")
                "text": text
            }
        }

    dataSend2 = {
    "version": "2.0",
    "template": {
        "outputs": [
        {
            "carousel": {
            "type": "basicCard",
            "items": [
                {
                "title": "보물상자",
                "description": "보물상자 안에는 뭐가 있을까",
                "thumbnail": {
                    "imageUrl": "http://k.kakaocdn.net/dn/83BvP/bl20duRC1Q1/lj3JUcmrzC53YIjNDkqbWK/i_6piz1p.jpg"
                },
                "buttons": [
                    {
                    "action": "message",
                    "label": "열어보기",
                    "messageText": "짜잔! 우리가 찾던 보물입니다"
                    },
                    {
                    "action":  "webLink",
                    "label": "구경하기",
                    "webLinkUrl": "https://e.kakao.com/t/hello-ryan"
                    }
                ]
                },
                {
                "title": "보물상자2",
                "description": "보물상자2 안에는 뭐가 있을까",
                "thumbnail": {
                    "imageUrl": "http://k.kakaocdn.net/dn/83BvP/bl20duRC1Q1/lj3JUcmrzC53YIjNDkqbWK/i_6piz1p.jpg"
                },
                "buttons": [
                    {
                    "action": "message",
                    "label": "열어보기",
                    "messageText": "짜잔! 우리가 찾던 보물입니다"
                    },
                    {
                    "action":  "webLink",
                    "label": "구경하기",
                    "webLinkUrl": "https://e.kakao.com/t/hello-ryan"
                    }
                ]
                },
                {
                "title": "보물상자3",
                "description": "보물상자3 안에는 뭐가 있을까",
                "thumbnail": {
                    "imageUrl": "http://k.kakaocdn.net/dn/83BvP/bl20duRC1Q1/lj3JUcmrzC53YIjNDkqbWK/i_6piz1p.jpg"
                },
                "buttons": [
                    {
                    "action": "message",
                    "label": "열어보기",
                    "messageText": "짜잔! 우리가 찾던 보물입니다"
                    },
                    {
                    "action":  "webLink",
                    "label": "구경하기",
                    "webLinkUrl": "https://e.kakao.com/t/hello-ryan"
                    }
                ]
                }
            ]
            }
        }
        ]
    }
    }
 
    return jsonify(dataSend)
 
 
 
if __name__ == "__main__":
    app.run(host='0.0.0.0', port = 5000)